import os

import discord
import openpyxl
import requests
from discord.ext import commands

url1 = 'https://covid19-japan-web-api.now.sh/api//v1/prefectures'


class CovidCog(commands.Cog):

    def __init__(self, bot):
        self.bot = bot

    @commands.group(pass_context=True)
    async def covid(self, ctx):
        if ctx.author.bot:
            return

        if ctx.invoked_subcommand is not None:
            return

        prefectures = [""]
        if len(ctx.message.content.split(' ')) > 1:
            prefectures = [pre for pre in ctx.message.content.split(' ')[1:]]

        msg = 'COVID-19 感染状況\n'
        r = requests.get(url1)
        json = r.json()
        if prefectures[0] == "":
            for i in range(47):
                msg += (str(json[i]['name_ja']) + '  計 ' + str(json[i]['cases']) + ' 人  死者 ' + str(json[i]['deaths']) + ' 人\n')
        else:
            for prefecture in prefectures:
                for i in range(47):
                    if prefecture.lower() in json[i]['name_en'].lower():
                        msg += (str(json[i]['name_ja']) + '  計 ' + str(json[i]['cases']) + ' 人  死者 ' + str(json[i]['deaths']) + ' 人\n')
                        break
                else:
                    msg += f'{prefecture}が見つかりませんでした\n'
        await ctx.send(msg)

    @covid.command(pass_context=True)
    async def excel(self, ctx):
        """COVID-19の感染状況をEXCELファイルで取得する

        :param ctx:
        :return:
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "COVID-19"
        ws['A1'] = "COVID-19 感染状況"
        ws['A2'] = "都道府県"
        ws['B2'] = "感染者数"
        ws['C2'] = "死者数"

        prefectures = []
        if len(ctx.message.content.split(' ')) > 1:
            prefectures = [pre for pre in ctx.message.content.split(' ')[2:]]

        r = requests.get(url1)
        json = r.json()

        data = []
        if len(prefectures) == 0:
            for i in range(47):
                data.append((json[i]['name_ja'], json[i]['cases'], json[i]['deaths']))
        else:
            for prefecture in prefectures:
                for i in range(47):
                    if prefecture.lower() in json[i]['name_en'].lower():
                        data.append((json[i]['name_ja'], json[i]['cases'], json[i]['deaths']))
                        break

        for i, x in enumerate(range(1, 4)):
            for j, y in enumerate(range(3, 3+len(data))):
                ws.cell(row=y, column=x, value=data[j][i])

        try:
            os.mkdir(os.path.join(self.bot.path, "tmp"))
        except:
            pass
        filepath = os.path.join(self.bot.path, "tmp", "covid19.xlsx")
        wb.save(filepath)

        await ctx.send(file=discord.File(filepath))


def setup(bot):
    bot.add_cog(CovidCog(bot))